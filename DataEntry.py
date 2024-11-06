import tkinter as tk
from tkinter import messagebox
from tkinter import ttk
import speech_recognition as sr
from openpyxl import Workbook, load_workbook
from datetime import datetime
import threading  # Import threading to run the speech recognition in the background

# Global variable to store the last entered date
previous_date = ""
selected_row_id = None  # This will store the ID of the selected row for updating

# Function to write data to Excel
def add_to_excel(date, record_id, name, price):
    global previous_date  # Use the global variable

    try:
        # Check if the Excel file exists, if not, create it with headers
        try:
            wb = load_workbook('data.xlsx')
            ws = wb.active
        except FileNotFoundError:
            wb = Workbook()
            ws = wb.active
            ws.append(["Date", "ID", "Name", "Price"])  # Adding headers in the first row

        # Adding the data to the next available row
        ws.append([date, record_id, name, price])
        wb.save('data.xlsx')

        messagebox.showinfo("Success", "Data added to Excel successfully!")
        clear_fields()
        update_table()  # Update the table view after adding data

        # Update the previous date with the current date
        previous_date = date
        entry_date.delete(0, tk.END)
        entry_date.insert(0, previous_date)  # Set the previous date back into the text field

    except Exception as e:
        messagebox.showerror("Error", f"Failed to add data: {e}")

# Function to update the existing row in Excel
def update_in_excel(date, record_id, name, price):
    global selected_row_id

    if selected_row_id is None:
        messagebox.showwarning("Selection Error", "No row selected for update!")
        return

    try:
        # Load the Excel file
        wb = load_workbook('data.xlsx')
        ws = wb.active

        # Find the row with the matching ID and update it
        for row in ws.iter_rows(min_row=2):  # Skip header row
            if row[1].value == selected_row_id:
                row[0].value = date  # Update the date
                row[1].value = record_id  # Update the ID
                row[2].value = name  # Update the name
                row[3].value = price  # Update the price
                break

        wb.save('data.xlsx')
        messagebox.showinfo("Success", "Data updated in Excel successfully!")
        clear_fields()
        update_table()  # Update the table view after editing

    except Exception as e:
        messagebox.showerror("Error", f"Failed to update data: {e}")

# Function to clear text fields
def clear_fields():
    entry_id.delete(0, tk.END)
    entry_name.delete(0, tk.END)
    entry_price.delete(0, tk.END)

# Function to update the table view
def update_table():
    # Clear the current table contents
    for row in tree.get_children():
        tree.delete(row)

    try:
        # Load the Excel file and read the data
        wb = load_workbook('data.xlsx')
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip the header row
            tree.insert("", tk.END, values=row)
    except Exception as e:
        messagebox.showerror("Error", f"Failed to load data: {e}")

# Function that is triggered when the button is clicked to add new data
def on_submit():
    date = entry_date.get()
    record_id = entry_id.get()
    name = entry_name.get()
    price = entry_price.get()

    if not date or not record_id or not name or not price:
        messagebox.showwarning("Input Error", "All fields must be filled!")
        return

    try:
        # Check if the date format is correct (DD/MM/YYYY)
        datetime.strptime(date, "%d/%m/%Y")
    except ValueError:
        messagebox.showwarning("Input Error", "Date must be in DD/MM/YYYY format!")
        return

    add_to_excel(date, record_id, name, price)

# Function that is triggered when a row in the table is selected
def on_row_select(event):
    global selected_row_id

    # Get the selected item (row)
    selected_item = tree.selection()
    if not selected_item:
        return

    # Get values from the selected row
    selected_values = tree.item(selected_item, 'values')
    
    # Set the entry fields with the selected row data
    entry_date.delete(0, tk.END)
    entry_date.insert(0, selected_values[0])  # Date
    entry_id.delete(0, tk.END)
    entry_id.insert(0, selected_values[1])  # ID
    entry_name.delete(0, tk.END)
    entry_name.insert(0, selected_values[2])  # Name
    entry_price.delete(0, tk.END)
    entry_price.insert(0, selected_values[3])  # Price

    # Set the selected row ID for updating
    selected_row_id = selected_values[1]  # Assuming ID is the second column

# Function to start voice recognition for name input in a separate thread
def listen_for_name():
    # Use threading to run the speech recognition in the background
    def recognize_name():
        recognizer = sr.Recognizer()
        microphone = sr.Microphone()

        with microphone as source:
            recognizer.adjust_for_ambient_noise(source)  # Adjust for ambient noise
            print("Listening for your name...")
            audio = recognizer.listen(source)

        try:
            # Recognize the speech and ask for confirmation
            name = recognizer.recognize_google(audio)
            print(f"Did you say '{name}'? (Yes/No)")

            def confirm_name():
                confirmation = messagebox.askquestion("Confirm Name", f"Did you say '{name}'?")
                if confirmation == 'yes':
                    entry_name.delete(0, tk.END)
                    entry_name.insert(0, name)  # Set the recognized name in the entry field
                else:
                    messagebox.showinfo("Retry", "Please say your name again.")
                    listen_for_name()  # If no, try again

            confirm_name()

        except sr.UnknownValueError:
            messagebox.showerror("Error", "Sorry, I couldn't understand the audio. Please try again.")
        except sr.RequestError:
            messagebox.showerror("Error", "Sorry, the speech service is unavailable.")

    # Start the recognize_name function in a separate thread
    threading.Thread(target=recognize_name, daemon=True).start()

# Set up the main window
root = tk.Tk()
root.title("Excel Data Entry")
root.geometry("900x400")  # Changed window geometry to 900x400

# Font for all widgets
font = ("Helvetica", 12)

# Create a frame for the form (left side)
frame_form = tk.Frame(root)
frame_form.grid(row=0, column=0, padx=20, pady=20)

# Label and entry for Date
tk.Label(frame_form, text="Date (DD/MM/YYYY)", font=font).grid(row=0, column=0, padx=10, pady=5)
entry_date = tk.Entry(frame_form, font=font)
entry_date.grid(row=0, column=1, padx=10, pady=5)

# If there is a previous date, set it in the date entry field
if previous_date:
    entry_date.insert(0, previous_date)

# Label and entry for ID
tk.Label(frame_form, text="ID", font=font).grid(row=1, column=0, padx=10, pady=5)
entry_id = tk.Entry(frame_form, font=font)
entry_id.grid(row=1, column=1, padx=10, pady=5)

# Label and entry for Name
tk.Label(frame_form, text="Name", font=font).grid(row=2, column=0, padx=10, pady=5)
entry_name = tk.Entry(frame_form, font=font)
entry_name.grid(row=2, column=1, padx=10, pady=5)

# Button for voice input next to the Name field
mic_button = tk.Button(frame_form, text="M", font=font, command=listen_for_name)
mic_button.grid(row=2, column=2, padx=10, pady=5)

# Label and entry for Price
tk.Label(frame_form, text="Price", font=font).grid(row=3, column=0, padx=10, pady=5)
entry_price = tk.Entry(frame_form, font=font)
entry_price.grid(row=3, column=1, padx=10, pady=5)

# Submit Button
submit_button = tk.Button(frame_form, text="Add to Excel", font=font, command=on_submit)
submit_button.grid(row=4, column=0, columnspan=3, pady=20)

# Update Button
update_button = tk.Button(frame_form, text="Update Excel", font=font, command=lambda: update_in_excel(entry_date.get(), entry_id.get(), entry_name.get(), entry_price.get()))
update_button.grid(row=5, column=0, columnspan=3, pady=10)

# Create a frame for the table (right side)
frame_table = tk.Frame(root)
frame_table.grid(row=0, column=1, padx=20, pady=20)

# Create the table using Treeview
tree = ttk.Treeview(frame_table, columns=("Date", "ID", "Name", "Price"), show="headings", selectmode="browse")
tree.heading("Date", text="Date", anchor=tk.W)
tree.heading("ID", text="ID", anchor=tk.W)
tree.heading("Name", text="Name", anchor=tk.W)
tree.heading("Price", text="Price", anchor=tk.W)

# Define column widths
tree.column("Date", width=100)
tree.column("ID", width=80)
tree.column("Name", width=150)
tree.column("Price", width=100)

# Apply font to the Treeview widget
tree.tag_configure("center", font=("Helvetica", 12))  # Adjust font for treeview

tree.grid(row=0, column=0)

# Bind row selection event
tree.bind("<ButtonRelease-1>", on_row_select)

# Initially load the data into the table
update_table()

# Start the tkinter event loop
root.mainloop()
